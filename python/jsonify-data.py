import openpyxl
import json
from datetime import datetime

def process_excel_to_json(excel_file_name, year):
    wb = openpyxl.load_workbook(excel_file_name)
    get_event_data(wb, year)
    get_player_data(wb, year, 'mpo')
    get_player_data(wb, year,'fpo')
        

def get_event_data(wb, year):
    ws = wb['events']
    json_path = '../lib/seeds/events.json'
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            date_range = row[0].replace('April', 'Apr').split(' - ')
            start_day = date_range[0].split()[1]
            end_day = date_range[1] 
            month = date_range[0].split()[0]
            start_day = start_day.zfill(2)
            end_day = end_day.zfill(2)
            event_start = f"{month} {start_day}, {year}"
            event_end = f"{month} {end_day}, {year}"
            event_start = datetime.strptime(event_start, '%b %d, %Y').strftime('%Y-%m-%d')
            event_end = datetime.strptime(event_end, '%b %d, %Y').strftime('%Y-%m-%d')
            events.append({
                "event_name": row[2],
                "event_date_start": event_start,
                "event_date_end": event_end,
                "location": row[3],
            })
        except Exception as e:
            print(f"An error occurred while processing row: {row}")
            print(e)
            # Skip this row and continue to the next
            continue
    with open(json_path, 'w') as f:
        json.dump(events, f, indent=4)
    print(f"Event data successfully written to {json_path}")


def get_player_data(wb, year, gender):
    ws = wb[gender]
    json_path = f'../lib/seeds/{gender}_players.json'
    players = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        player_name = row[0]
        pdga_number = row[1]
        players.append({
            "name": player_name,
            "pdga_number": pdga_number
        })
    with open(json_path, 'w') as f:
        json.dump(players, f, indent=4)
    print(f"{gender.upper()} player data successfully written to {json_path}")

def main():
    excel_file_name = 'dg-data-2024.xlsx'
    year = excel_file_name.split('-')[-1].split('.')[0]
    process_excel_to_json(excel_file_name, year)


if __name__ == "__main__":
    main()
